# %% [Notebook cell 1]
# Step 1: Load the Data

# Import required library
import pandas as pd

# Load ADCN dataset
adcn = pd.read_excel("/content/ADCN.xlsx")
print("ADCN Dataset Loaded Successfully!\n")

# Check first 5 rows
print("ADCN First 5 Rows:")
print(adcn.head(), "\n")

# Check data types and basic info
print("ADCN Info:")
print(adcn.info(), "\n")

# Remove leading/trailing spaces from column names
adcn.columns = adcn.columns.str.strip()

# Check missing values using correct names
print("ADCN Missing Values:")
print(adcn[['dn', 'de', 'du', 'Temp.Dry [°C]', 'RelHumidity [%]', 'Press.QFF [hPa]']].isnull().sum(), "\n")


# Load WAGN dataset
wagn = pd.read_excel("/content/WAGN  .xlsx")
print("WAGN Dataset Loaded Successfully!\n")

# Check first 5 rows
print("WAGN First 5 Rows:")
print(wagn.head(), "\n")

# Check data types and basic info
print("WAGN Info:")
print(wagn.info(), "\n")

# Strip any leading/trailing spaces just in case
wagn.columns = wagn.columns.str.strip()

# Check missing values using correct WAGN column names
print("WAGN Missing Values:")
print(wagn[['d North (mm)', 'd East (mm)', 'd Up (mm)',
            'Temp.Dry [°C]', 'RelHumidity [%]', 'Press.QFF [hPa]']].isnull().sum())

# %% [Notebook cell 3]
# Step 2: Data Cleaning and Preparation

# 1. Harmonize column names for WAGN
wagn.rename(columns={'d North (mm)': 'dn', 'd East (mm)': 'de', 'd Up (mm)': 'du'}, inplace=True)

# 2. Ensure Date column exists for WAGN
wagn['Date'] = pd.to_datetime(wagn['DOY'], format='%Y%m%d')

# 3. Add Month column
adcn['Month'] = adcn['Date'].dt.month
wagn['Month'] = wagn['Date'].dt.month

# 4. Add Season column
def get_season(month):
    if month in [12, 1, 2, 3]:
        return 'Winter'
    elif month in [6, 7, 8, 9]:
        return 'Summer'
    elif month in [3, 4, 5]:
        return 'Spring'
    else:
        return 'Autumn'

adcn['Season'] = adcn['Month'].apply(get_season)
wagn['Season'] = wagn['Month'].apply(get_season)

# Quick check
print("ADCN Head with Month & Season:")
print(adcn.head(), "\n")
print("WAGN Head with Month & Season:")
print(wagn.head())

# %% [Notebook cell 5]
# Step 3: Plot GNSS Errors over Time

import matplotlib.pyplot as plt

# Function to plot dn, de, du
def plot_gnss_errors(df, station_name):
    plt.figure(figsize=(14,6))

    plt.plot(df['Date'], df['dn'], label='dN (mm)', color='blue')
    plt.plot(df['Date'], df['de'], label='dE (mm)', color='green')
    plt.plot(df['Date'], df['du'], label='dUp (mm)', color='red')

    plt.title(f'{station_name} GNSS Errors Over Time')
    plt.xlabel('Date')
    plt.ylabel('Error (mm)')
    plt.legend()
    plt.grid(True)
    plt.show()

# Plot for ADCN
plot_gnss_errors(adcn, 'ADCN')

# Plot for WAGN
plot_gnss_errors(wagn, 'WAGN')

# %% [Notebook cell 7]
# Step 4: Check for Outliers

def plot_outliers(df, station_name):
    plt.figure(figsize=(12,6))

    # Boxplots for dn, de, du
    plt.boxplot([df['dn'], df['de'], df['du']], labels=['dN', 'dE', 'dUp'])
    plt.title(f'{station_name} GNSS Errors - Outlier Check')
    plt.ylabel('Error (mm)')
    plt.grid(True)
    plt.show()

# ADCN outliers
plot_outliers(adcn, 'ADCN')

# WAGN outliers
plot_outliers(wagn, 'WAGN')

# %% [Notebook cell 9]
# Step 5: Max/Min and Daily/Weekly/Monthly Errors

# Import needed library
import numpy as np

# Function to compute statistics
def compute_error_stats(df, station_name):
    # Daily max/min
    daily_stats = df.groupby('Date')[['dn','de','du']].agg([np.max, np.min, np.mean])
    print(f"--- {station_name} Daily Max/Min/Mean ---")
    print(daily_stats.head(), "\n")

    # Weekly max/min (week number)
    df['Week'] = df['Date'].dt.isocalendar().week
    weekly_stats = df.groupby('Week')[['dn','de','du']].agg([np.max, np.min, np.mean])
    print(f"--- {station_name} Weekly Max/Min/Mean ---")
    print(weekly_stats.head(), "\n")

    # Monthly max/min
    monthly_stats = df.groupby('Month')[['dn','de','du']].agg([np.max, np.min, np.mean])
    print(f"--- {station_name} Monthly Max/Min/Mean ---")
    print(monthly_stats.head(), "\n")

    # Identify component with maximum error
    max_errors = df[['dn','de','du']].abs().max()
    max_component = max_errors.idxmax()
    max_value = max_errors.max()
    print(f"{station_name} Maximum Absolute Error: {max_value} mm in component {max_component}\n")

    return daily_stats, weekly_stats, monthly_stats

# ADCN stats
adcn_daily, adcn_weekly, adcn_monthly = compute_error_stats(adcn, 'ADCN')

# WAGN stats
wagn_daily, wagn_weekly, wagn_monthly = compute_error_stats(wagn, 'WAGN')

# %% [Notebook cell 10]
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# -----------------------------
# 1️⃣ Daily, Weekly, Monthly Stats
# -----------------------------

# Daily
daily_stats_adcn = adcn.groupby('Date')[['dn','de','du']].agg(['max','min','mean'])
daily_stats_wagn = wagn.groupby('Date')[['dn','de','du']].agg(['max','min','mean'])

# Weekly
weekly_stats_adcn = adcn.groupby('Week')[['dn','de','du']].agg(['max','min','mean'])
weekly_stats_wagn = wagn.groupby('Week')[['dn','de','du']].agg(['max','min','mean'])

# Monthly
monthly_stats_adcn = adcn.groupby('Month')[['dn','de','du']].agg(['max','min','mean'])
monthly_stats_wagn = wagn.groupby('Month')[['dn','de','du']].agg(['max','min','mean'])

# -----------------------------
# 2️⃣ Maximum Absolute Error
# -----------------------------
max_error_adcn = adcn[['dn','de','du']].abs().max().max()
max_error_wagn = wagn[['dn','de','du']].abs().max().max()

print("ADCN Maximum Absolute Error:", max_error_adcn)
print("WAGN Maximum Absolute Error:", max_error_wagn)

# -----------------------------
# 3️⃣ Seasonal Stats
# -----------------------------
season_stats_adcn = adcn.groupby('Season')[['dn','de','du']].agg(['max','min','mean'])
season_stats_wagn = wagn.groupby('Season')[['dn','de','du']].agg(['max','min','mean'])

print("ADCN Seasonal Stats:\n", season_stats_adcn)
print("WAGN Seasonal Stats:\n", season_stats_wagn)

# -----------------------------
# 4️⃣ Identify Max Error Day/Week/Month per Component
# -----------------------------
def max_error_period(stats_df, period_type):
    components = ['dn','de','du']
    max_days = {}
    for comp in components:
        max_val = stats_df[comp]['max'].max()
        period = stats_df[comp]['max'].idxmax()
        max_days[comp] = (period, max_val)
    print(f"\nMaximum Error by {period_type}: {max_days}")

max_error_period(daily_stats_adcn, "Day (ADCN)")
max_error_period(weekly_stats_adcn, "Week (ADCN)")
max_error_period(monthly_stats_adcn, "Month (ADCN)")

max_error_period(daily_stats_wagn, "Day (WAGN)")
max_error_period(weekly_stats_wagn, "Week (WAGN)")
max_error_period(monthly_stats_wagn, "Month (WAGN)")

# -----------------------------
# 5️⃣ Plot Max, Min, Mean per Component
# -----------------------------
def plot_error_stats(stats_df, title):
    plt.figure(figsize=(12,5))
    for comp in ['dn','de','du']:
        plt.plot(stats_df.index, stats_df[comp]['max'], label=f'{comp} Max')
        plt.plot(stats_df.index, stats_df[comp]['min'], label=f'{comp} Min')
        plt.plot(stats_df.index, stats_df[comp]['mean'], label=f'{comp} Mean', linestyle='--')
    plt.title(title)
    plt.xlabel('Period')
    plt.ylabel('Error (mm)')
    plt.legend()
    plt.grid(True)
    plt.show()

plot_error_stats(daily_stats_adcn, "ADCN Daily Errors")
plot_error_stats(daily_stats_wagn, "WAGN Daily Errors")
plot_error_stats(monthly_stats_adcn, "ADCN Monthly Errors")
plot_error_stats(monthly_stats_wagn, "WAGN Monthly Errors")
plot_error_stats(season_stats_adcn, "ADCN Seasonal Errors")
plot_error_stats(season_stats_wagn, "WAGN Seasonal Errors")

# %% [Notebook cell 11]
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# -----------------------------
# 📊 Export all tables to Excel
# -----------------------------
output_file = "GNSS_Analysis_Results.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    # ADCN
    daily_stats_adcn.to_excel(writer, sheet_name="ADCN_Daily")
    weekly_stats_adcn.to_excel(writer, sheet_name="ADCN_Weekly")
    monthly_stats_adcn.to_excel(writer, sheet_name="ADCN_Monthly")
    season_stats_adcn.to_excel(writer, sheet_name="ADCN_Seasonal")

    # WAGN
    daily_stats_wagn.to_excel(writer, sheet_name="WAGN_Daily")
    weekly_stats_wagn.to_excel(writer, sheet_name="WAGN_Weekly")
    monthly_stats_wagn.to_excel(writer, sheet_name="WAGN_Monthly")
    season_stats_wagn.to_excel(writer, sheet_name="WAGN_Seasonal")

    # Max error summary
    max_summary = pd.DataFrame({
        "Station": ["ADCN", "WAGN"],
        "Max Absolute Error (mm)": [max_error_adcn, max_error_wagn]
    })
    max_summary.to_excel(writer, sheet_name="Max_Errors", index=False)

print(f"✅ Tables saved to {output_file}")

# -----------------------------
# 🖼 Save plots as images
# -----------------------------
def save_error_plot(stats_df, title, filename):
    plt.figure(figsize=(12,5))
    for comp in ['dn','de','du']:
        plt.plot(stats_df.index, stats_df[comp]['max'], label=f'{comp} Max')
        plt.plot(stats_df.index, stats_df[comp]['min'], label=f'{comp} Min')
        plt.plot(stats_df.index, stats_df[comp]['mean'], label=f'{comp} Mean', linestyle='--')
    plt.title(title)
    plt.xlabel('Period')
    plt.ylabel('Error (mm)')
    plt.legend()
    plt.grid(True)
    plt.savefig(filename, dpi=300, bbox_inches="tight")
    plt.close()

# Save a few plots
save_error_plot(daily_stats_adcn, "ADCN Daily Errors", "ADCN_Daily.png")
save_error_plot(daily_stats_wagn, "WAGN Daily Errors", "WAGN_Daily.png")
save_error_plot(monthly_stats_adcn, "ADCN Monthly Errors", "ADCN_Monthly.png")
save_error_plot(monthly_stats_wagn, "WAGN Monthly Errors", "WAGN_Monthly.png")
save_error_plot(season_stats_adcn, "ADCN Seasonal Errors", "ADCN_Seasonal.png")
save_error_plot(season_stats_wagn, "WAGN Seasonal Errors", "WAGN_Seasonal.png")

# -----------------------------
# 📎 Insert plots into Excel
# -----------------------------
wb = load_workbook(output_file)

def insert_plot(sheet_name, image_file):
    ws = wb.create_sheet(sheet_name)
    img = Image(image_file)
    ws.add_image(img, "A1")

insert_plot("ADCN_Daily_Plot", "ADCN_Daily.png")
insert_plot("WAGN_Daily_Plot", "WAGN_Daily.png")
insert_plot("ADCN_Monthly_Plot", "ADCN_Monthly.png")
insert_plot("WAGN_Monthly_Plot", "WAGN_Monthly.png")
insert_plot("ADCN_Seasonal_Plot", "ADCN_Seasonal.png")
insert_plot("WAGN_Seasonal_Plot", "WAGN_Seasonal.png")

# Save workbook
wb.save(output_file)
print(f"✅ All plots inserted into {output_file}")

# %% [Notebook cell 13]
# 1️⃣ Seasonal Statistics (ADCN & WAGN)

def seasonal_stats(df):
    stats = df.groupby('Season')[['dn','de','du']].agg(['mean','max','min','std'])
    stats['dn_variation'] = stats['dn']['std']
    stats['de_variation'] = stats['de']['std']
    stats['du_variation'] = stats['du']['std']
    return stats

season_stats_adcn = seasonal_stats(adcn)
season_stats_wagn = seasonal_stats(wagn)

print("ADCN Seasonal Stats:\n", season_stats_adcn)
print("\nWAGN Seasonal Stats:\n", season_stats_wagn)


# 2️⃣ Identify Representative High-Error Days

def representative_days(df, season_name):
    # Filter by season
    season_df = df[df['Season']==season_name]
    # Compute daily absolute error sum (dn + de + du)
    season_df['abs_total_error'] = season_df[['dn','de','du']].abs().sum(axis=1)
    # Pick day with highest total error
    high_error_day = season_df.loc[season_df['abs_total_error'].idxmax()]
    return high_error_day['Date'], high_error_day['abs_total_error']

# ADCN representative days
winter_day_adcn, winter_error_adcn = representative_days(adcn, 'Winter')
summer_day_adcn, summer_error_adcn = representative_days(adcn, 'Summer')

# WAGN representative days
winter_day_wagn, winter_error_wagn = representative_days(wagn, 'Winter')
summer_day_wagn, summer_error_wagn = representative_days(wagn, 'Summer')

print(f"\nADCN Representative Days:")
print(f"Winter High-Error Day: {winter_day_adcn} | Total Error: {winter_error_adcn:.2f} mm")
print(f"Summer High-Error Day: {summer_day_adcn} | Total Error: {summer_error_adcn:.2f} mm")

print(f"\nWAGN Representative Days:")
print(f"Winter High-Error Day: {winter_day_wagn} | Total Error: {winter_error_wagn:.2f} mm")
print(f"Summer High-Error Day: {summer_day_wagn} | Total Error: {summer_error_wagn:.2f} mm")

# %% [Notebook cell 14]
import pandas as pd
from openpyxl import load_workbook

output_file = "GNSS_Seasonal_Analysis.xlsx"

# -----------------------------
# 1️⃣ Save Seasonal Statistics
# -----------------------------
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    season_stats_adcn.to_excel(writer, sheet_name="ADCN_Seasonal_Stats")
    season_stats_wagn.to_excel(writer, sheet_name="WAGN_Seasonal_Stats")

# -----------------------------
# 2️⃣ Save Representative High-Error Days
# -----------------------------
rep_days_summary = pd.DataFrame({
    "Station": ["ADCN", "ADCN", "WAGN", "WAGN"],
    "Season": ["Winter", "Summer", "Winter", "Summer"],
    "High_Error_Day": [
        winter_day_adcn, summer_day_adcn,
        winter_day_wagn, summer_day_wagn
    ],
    "Total_Error_mm": [
        winter_error_adcn, summer_error_adcn,
        winter_error_wagn, summer_error_wagn
    ]
})

# Append representative days into Excel
with pd.ExcelWriter(output_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
    rep_days_summary.to_excel(writer, sheet_name="Representative_Days", index=False)

print(f"✅ Seasonal stats + representative days saved in {output_file}")

# %% [Notebook cell 15]
import pandas as pd
import matplotlib.pyplot as plt

# Assuming your datasets already have 'Date' column as datetime
# And high-error days from Step 6
adcn_winter_day = pd.Timestamp('2024-03-14')
adcn_summer_day = pd.Timestamp('2024-06-22')

wagn_winter_day = pd.Timestamp('2024-12-26')
wagn_summer_day = pd.Timestamp('2024-08-20')

# Function to extract hourly data and plot errors
def plot_hourly_errors(df, day, station_name):
    # Filter the day
    day_data = df[df['Date'] == day].copy()

    # If you have actual hourly timestamps, use them; else simulate hours
    # Here we assume one row per hour or evenly spaced measurements
    day_data['Hour'] = range(len(day_data))

    plt.figure(figsize=(12,6))
    plt.plot(day_data['Hour'], day_data['dn'], label='dN', marker='o')
    plt.plot(day_data['Hour'], day_data['de'], label='dE', marker='s')
    plt.plot(day_data['Hour'], day_data['du'], label='dUp', marker='^')

    plt.title(f'{station_name} GNSS Errors on {day.date()} (Hourly)')
    plt.xlabel('Hour')
    plt.ylabel('Error (mm)')
    plt.grid(True)
    plt.legend()
    plt.show()

    # Optional: compute hourly statistics
    stats = day_data[['dn','de','du']].agg(['mean','std','max','min'])
    print(f'Hourly Statistics for {station_name} on {day.date()}:\n', stats)

# ADCN Hourly Plots
plot_hourly_errors(adcn, adcn_winter_day, 'ADCN - Winter High Error')
plot_hourly_errors(adcn, adcn_summer_day, 'ADCN - Summer High Error')

# WAGN Hourly Plots
plot_hourly_errors(wagn, wagn_winter_day, 'WAGN - Winter High Error')
plot_hourly_errors(wagn, wagn_summer_day, 'WAGN - Summer High Error')

# %% [Notebook cell 16]
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

output_file = "GNSS_Hourly_Analysis.xlsx"

# Function to extract hourly data, plot, and return stats
def plot_hourly_errors(df, day, station_name, filename_prefix):
    # Filter the day
    day_data = df[df['Date'] == day].copy()
    day_data['Hour'] = range(len(day_data))

    # Plot
    plt.figure(figsize=(12,6))
    plt.plot(day_data['Hour'], day_data['dn'], label='dN', marker='o')
    plt.plot(day_data['Hour'], day_data['de'], label='dE', marker='s')
    plt.plot(day_data['Hour'], day_data['du'], label='dUp', marker='^')

    plt.title(f'{station_name} GNSS Errors on {day.date()} (Hourly)')
    plt.xlabel('Hour')
    plt.ylabel('Error (mm)')
    plt.grid(True)
    plt.legend()
    plot_file = f"{filename_prefix}_{day.date()}.png"
    plt.savefig(plot_file, dpi=300, bbox_inches="tight")
    plt.close()

    # Hourly statistics
    stats = day_data[['dn','de','du']].agg(['mean','std','max','min'])
    stats['Day'] = day.date()
    stats['Station'] = station_name
    return stats, plot_file

# -----------------------------
# Generate stats + plots
# -----------------------------
results = []

adcn_winter_stats, adcn_winter_plot = plot_hourly_errors(adcn, adcn_winter_day, 'ADCN - Winter High Error', "ADCN_Winter")
adcn_summer_stats, adcn_summer_plot = plot_hourly_errors(adcn, adcn_summer_day, 'ADCN - Summer High Error', "ADCN_Summer")

wagn_winter_stats, wagn_winter_plot = plot_hourly_errors(wagn, wagn_winter_day, 'WAGN - Winter High Error', "WAGN_Winter")
wagn_summer_stats, wagn_summer_plot = plot_hourly_errors(wagn, wagn_summer_day, 'WAGN - Summer High Error', "WAGN_Summer")

# Combine results
hourly_summary = pd.concat([adcn_winter_stats, adcn_summer_stats, wagn_winter_stats, wagn_summer_stats])

# -----------------------------
# Save to Excel
# -----------------------------
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    hourly_summary.to_excel(writer, sheet_name="Hourly_Stats")

# Insert plots
wb = load_workbook(output_file)

def insert_plot(sheet_name, image_file):
    ws = wb.create_sheet(sheet_name)
    img = Image(image_file)
    ws.add_image(img, "A1")

insert_plot("ADCN_Winter_Plot", adcn_winter_plot)
insert_plot("ADCN_Summer_Plot", adcn_summer_plot)
insert_plot("WAGN_Winter_Plot", wagn_winter_plot)
insert_plot("WAGN_Summer_Plot", wagn_summer_plot)

wb.save(output_file)

print(f"✅ Hourly stats + plots saved in {output_file}")

# %% [Notebook cell 18]
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# Ensure your datasets are loaded as: adcn, wagn
# Components to analyze
components = ['dn', 'de', 'du']

# Function to compute stats
def compute_stats(df, components):
    stats = pd.DataFrame(index=components, columns=['Mean','SD','RMS','MAE','RMSE'])
    for comp in components:
        stats.loc[comp,'Mean'] = df[comp].mean()
        stats.loc[comp,'SD'] = df[comp].std()
        stats.loc[comp,'RMS'] = np.sqrt((df[comp]**2).mean())
        stats.loc[comp,'MAE'] = df[comp].abs().mean()
        stats.loc[comp,'RMSE'] = np.sqrt(((df[comp])**2).mean())
    return stats.astype(float)

# Compute stats for both datasets
adcn_stats = compute_stats(adcn, components)
wagn_stats = compute_stats(wagn, components)

print("ADCN Statistical Metrics:\n", adcn_stats)
print("\nWAGN Statistical Metrics:\n", wagn_stats)

# Plotting comparison
metrics = ['Mean','SD','RMS','MAE','RMSE']
x = np.arange(len(components))
width = 0.35  # bar width

fig, axes = plt.subplots(1, len(metrics), figsize=(20,5), constrained_layout=True)

for i, metric in enumerate(metrics):
    axes[i].bar(x - width/2, adcn_stats[metric], width, label='ADCN', color='skyblue')
    axes[i].bar(x + width/2, wagn_stats[metric], width, label='WAGN', color='orange')
    axes[i].set_xticks(x)
    axes[i].set_xticklabels(components)
    axes[i].set_title(metric)
    axes[i].set_ylabel('mm')
    axes[i].legend()

plt.suptitle('GNSS Error Statistical Comparison: ADCN vs WAGN')
plt.show()

# %% [Notebook cell 19]
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

output_file = "GNSS_Stats_Comparison.xlsx"

# -----------------------------
# Function to compute stats
# -----------------------------
def compute_stats(df, components):
    stats = pd.DataFrame(index=components, columns=['Mean','SD','RMS','MAE','RMSE'])
    for comp in components:
        stats.loc[comp,'Mean'] = df[comp].mean()
        stats.loc[comp,'SD'] = df[comp].std()
        stats.loc[comp,'RMS'] = np.sqrt((df[comp]**2).mean())
        stats.loc[comp,'MAE'] = df[comp].abs().mean()
        stats.loc[comp,'RMSE'] = np.sqrt(((df[comp])**2).mean())
    return stats.astype(float)

# -----------------------------
# Compute stats for both datasets
# -----------------------------
components = ['dn', 'de', 'du']
adcn_stats = compute_stats(adcn, components)
wagn_stats = compute_stats(wagn, components)

# Combine into one dataframe with station labels
combined_stats = pd.concat({"ADCN": adcn_stats, "WAGN": wagn_stats})

print("ADCN Statistical Metrics:\n", adcn_stats)
print("\nWAGN Statistical Metrics:\n", wagn_stats)

# -----------------------------
# Plotting comparison
# -----------------------------
metrics = ['Mean','SD','RMS','MAE','RMSE']
x = np.arange(len(components))
width = 0.35  # bar width

fig, axes = plt.subplots(1, len(metrics), figsize=(20,5), constrained_layout=True)

for i, metric in enumerate(metrics):
    axes[i].bar(x - width/2, adcn_stats[metric], width, label='ADCN', color='skyblue')
    axes[i].bar(x + width/2, wagn_stats[metric], width, label='WAGN', color='orange')
    axes[i].set_xticks(x)
    axes[i].set_xticklabels(components)
    axes[i].set_title(metric)
    axes[i].set_ylabel('mm')
    axes[i].legend()

plt.suptitle('GNSS Error Statistical Comparison: ADCN vs WAGN')
plot_file = "GNSS_Stats_Comparison.png"
plt.savefig(plot_file, dpi=300, bbox_inches="tight")
plt.close()

# -----------------------------
# Save to Excel
# -----------------------------
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    adcn_stats.to_excel(writer, sheet_name="ADCN_Stats")
    wagn_stats.to_excel(writer, sheet_name="WAGN_Stats")
    combined_stats.to_excel(writer, sheet_name="Combined")

# Insert plot into Excel
wb = load_workbook(output_file)
ws = wb.create_sheet("Comparison_Plot")
img = Image(plot_file)
ws.add_image(img, "A1")
wb.save(output_file)

print(f"✅ Results saved to {output_file} (stats + plots)")

# %% [Notebook cell 21]
import numpy as np
import matplotlib.pyplot as plt

# Function to compute 2D and 3D errors
def compute_2d_3d_errors(df):
    df = df.copy()
    df['Error_2D'] = np.sqrt(df['dn']**2 + df['de']**2)
    df['Error_3D'] = np.sqrt(df['dn']**2 + df['de']**2 + df['du']**2)
    return df

# Compute errors for ADCN and WAGN
adcn = compute_2d_3d_errors(adcn)
wagn = compute_2d_3d_errors(wagn)

# Plot 2D & 3D errors over time
def plot_errors_over_time(df, station_name):
    plt.figure(figsize=(12,6))
    plt.plot(df['Date'], df['Error_2D'], label='2D Error', color='blue')
    plt.plot(df['Date'], df['Error_3D'], label='3D Error', color='red')
    plt.title(f'{station_name} GNSS 2D & 3D Errors Over Time')
    plt.xlabel('Date')
    plt.ylabel('Error (mm)')
    plt.grid(True)
    plt.legend()
    plt.show()

plot_errors_over_time(adcn, 'ADCN')
plot_errors_over_time(wagn, 'WAGN')

# Compute seasonal statistics
def seasonal_error_stats(df):
    stats = df.groupby('Season')[['Error_2D','Error_3D']].agg(['mean','max','std','min','median'])
    return stats

adcn_seasonal_stats = seasonal_error_stats(adcn)
wagn_seasonal_stats = seasonal_error_stats(wagn)

print("ADCN Seasonal 2D/3D Error Stats:\n", adcn_seasonal_stats)
print("\nWAGN Seasonal 2D/3D Error Stats:\n", wagn_seasonal_stats)

# Optional: plot seasonal error comparison
def plot_seasonal_errors(df, station_name):
    seasonal_means = df.groupby('Season')[['Error_2D','Error_3D']].mean()
    seasonal_means.plot(kind='bar', figsize=(10,6))
    plt.title(f'{station_name} Average Seasonal 2D & 3D Errors')
    plt.ylabel('Mean Error (mm)')
    plt.grid(axis='y')
    plt.show()

plot_seasonal_errors(adcn, 'ADCN')
plot_seasonal_errors(wagn, 'WAGN')

# %% [Notebook cell 22]
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

output_file = "GNSS_2D3D_Error_Analysis.xlsx"

# -----------------------------
# Function to compute 2D and 3D errors
# -----------------------------
def compute_2d_3d_errors(df):
    df = df.copy()
    df['Error_2D'] = np.sqrt(df['dn']**2 + df['de']**2)
    df['Error_3D'] = np.sqrt(df['dn']**2 + df['de']**2 + df['du']**2)
    return df

# Apply
adcn = compute_2d_3d_errors(adcn)
wagn = compute_2d_3d_errors(wagn)

# -----------------------------
# Plot 2D & 3D errors over time
# -----------------------------
def plot_errors_over_time(df, station_name):
    plt.figure(figsize=(12,6))
    plt.plot(df['Date'], df['Error_2D'], label='2D Error', color='blue')
    plt.plot(df['Date'], df['Error_3D'], label='3D Error', color='red')
    plt.title(f'{station_name} GNSS 2D & 3D Errors Over Time')
    plt.xlabel('Date')
    plt.ylabel('Error (mm)')
    plt.grid(True)
    plt.legend()
    fname = f"{station_name}_2D3D_Errors.png"
    plt.savefig(fname, dpi=300, bbox_inches="tight")
    plt.close()
    return fname

adcn_time_plot = plot_errors_over_time(adcn, 'ADCN')
wagn_time_plot = plot_errors_over_time(wagn, 'WAGN')

# -----------------------------
# Compute seasonal statistics
# -----------------------------
def seasonal_error_stats(df):
    return df.groupby('Season')[['Error_2D','Error_3D']].agg(['mean','max','std','min','median'])

adcn_seasonal_stats = seasonal_error_stats(adcn)
wagn_seasonal_stats = seasonal_error_stats(wagn)

print("ADCN Seasonal 2D/3D Error Stats:\n", adcn_seasonal_stats)
print("\nWAGN Seasonal 2D/3D Error Stats:\n", wagn_seasonal_stats)

# -----------------------------
# Plot seasonal error comparison
# -----------------------------
def plot_seasonal_errors(df, station_name):
    seasonal_means = df.groupby('Season')[['Error_2D','Error_3D']].mean()
    ax = seasonal_means.plot(kind='bar', figsize=(10,6))
    plt.title(f'{station_name} Average Seasonal 2D & 3D Errors')
    plt.ylabel('Mean Error (mm)')
    plt.grid(axis='y')
    fname = f"{station_name}_Seasonal_Errors.png"
    plt.savefig(fname, dpi=300, bbox_inches="tight")
    plt.close()
    return fname

adcn_seasonal_plot = plot_seasonal_errors(adcn, 'ADCN')
wagn_seasonal_plot = plot_seasonal_errors(wagn, 'WAGN')

# -----------------------------
# Save results to Excel
# -----------------------------
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    adcn_seasonal_stats.to_excel(writer, sheet_name="ADCN_Seasonal_Stats")
    wagn_seasonal_stats.to_excel(writer, sheet_name="WAGN_Seasonal_Stats")

# Insert plots into Excel
wb = load_workbook(output_file)

# ADCN Plots
ws1 = wb.create_sheet("ADCN_Plots")
ws1.add_image(Image(adcn_time_plot), "A1")
ws1.add_image(Image(adcn_seasonal_plot), "A30")

# WAGN Plots
ws2 = wb.create_sheet("WAGN_Plots")
ws2.add_image(Image(wagn_time_plot), "A1")
ws2.add_image(Image(wagn_seasonal_plot), "A30")

wb.save(output_file)

print(f"✅ Results exported to {output_file} (stats + plots)")

# %% [Notebook cell 23]
# Step 8 Extended: Compute 2D & 3D Errors Daily, Weekly, Monthly

def compute_error_stats(df, freq='D'):
    """
    freq = 'D' for daily, 'W' for weekly, 'M' for monthly
    Returns max, min, mean, std for 2D and 3D errors
    """
    if freq == 'D':
        group_label = df['Date']
    elif freq == 'W':
        group_label = df['Date'].dt.isocalendar().week
    elif freq == 'M':
        group_label = df['Date'].dt.month
    else:
        raise ValueError("freq must be 'D', 'W', or 'M'")

    stats = df.groupby(group_label)[['Error_2D','Error_3D']].agg(['max','min','mean','std'])
    return stats

# Compute daily, weekly, monthly stats for ADCN
adcn_daily_stats = compute_error_stats(adcn, 'D')
adcn_weekly_stats = compute_error_stats(adcn, 'W')
adcn_monthly_stats = compute_error_stats(adcn, 'M')

# Compute daily, weekly, monthly stats for WAGN
wagn_daily_stats = compute_error_stats(wagn, 'D')
wagn_weekly_stats = compute_error_stats(wagn, 'W')
wagn_monthly_stats = compute_error_stats(wagn, 'M')

# Display summary
print("=== ADCN Daily 2D/3D Error Stats ===")
print(adcn_daily_stats.head())

print("\n=== ADCN Weekly 2D/3D Error Stats ===")
print(adcn_weekly_stats.head())

print("\n=== ADCN Monthly 2D/3D Error Stats ===")
print(adcn_monthly_stats.head())

print("\n=== WAGN Daily 2D/3D Error Stats ===")
print(wagn_daily_stats.head())

print("\n=== WAGN Weekly 2D/3D Error Stats ===")
print(wagn_weekly_stats.head())

print("\n=== WAGN Monthly 2D/3D Error Stats ===")
print(wagn_monthly_stats.head())

# %% [Notebook cell 24]
import pandas as pd
from openpyxl import load_workbook

output_file = "GNSS_2D3D_Error_Stats.xlsx"

# -----------------------------
# Step 8 Extended: Compute Error Stats
# -----------------------------
def compute_error_stats(df, freq='D'):
    """
    freq = 'D' for daily, 'W' for weekly, 'M' for monthly
    Returns max, min, mean, std for 2D and 3D errors
    """
    if freq == 'D':
        group_label = df['Date']
    elif freq == 'W':
        group_label = df['Date'].dt.isocalendar().week
    elif freq == 'M':
        group_label = df['Date'].dt.month
    else:
        raise ValueError("freq must be 'D', 'W', or 'M'")

    stats = df.groupby(group_label)[['Error_2D','Error_3D']].agg(['max','min','mean','std'])
    return stats

# Compute for ADCN
adcn_daily_stats = compute_error_stats(adcn, 'D')
adcn_weekly_stats = compute_error_stats(adcn, 'W')
adcn_monthly_stats = compute_error_stats(adcn, 'M')

# Compute for WAGN
wagn_daily_stats = compute_error_stats(wagn, 'D')
wagn_weekly_stats = compute_error_stats(wagn, 'W')
wagn_monthly_stats = compute_error_stats(wagn, 'M')

# -----------------------------
# Save Results to Excel
# -----------------------------
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    adcn_daily_stats.to_excel(writer, sheet_name="ADCN_Daily_2D3D")
    adcn_weekly_stats.to_excel(writer, sheet_name="ADCN_Weekly_2D3D")
    adcn_monthly_stats.to_excel(writer, sheet_name="ADCN_Monthly_2D3D")

    wagn_daily_stats.to_excel(writer, sheet_name="WAGN_Daily_2D3D")
    wagn_weekly_stats.to_excel(writer, sheet_name="WAGN_Weekly_2D3D")
    wagn_monthly_stats.to_excel(writer, sheet_name="WAGN_Monthly_2D3D")

print(f"✅ 2D/3D daily, weekly, monthly stats exported to {output_file}")

# %% [Notebook cell 25]
import matplotlib.pyplot as plt

def plot_error_stats(stats, title_prefix):
    """
    Plots 2D and 3D error stats (mean, max, min) for given stats DataFrame
    """
    for error_type in ['Error_2D','Error_3D']:
        plt.figure(figsize=(12,6))
        plt.plot(stats[error_type]['mean'], label='Mean', marker='o')
        plt.plot(stats[error_type]['max'], label='Max', marker='x')
        plt.plot(stats[error_type]['min'], label='Min', marker='.')
        plt.title(f'{title_prefix} - {error_type} Error')
        plt.xlabel('Time Group')
        plt.ylabel('Error (mm)')
        plt.legend()
        plt.grid(True)
        plt.show()

# ----- ADCN Plots -----
print("ADCN Daily 2D/3D Errors")
plot_error_stats(adcn_daily_stats, 'ADCN Daily')

print("ADCN Weekly 2D/3D Errors")
plot_error_stats(adcn_weekly_stats, 'ADCN Weekly')

print("ADCN Monthly 2D/3D Errors")
plot_error_stats(adcn_monthly_stats, 'ADCN Monthly')

# ----- WAGN Plots -----
print("WAGN Daily 2D/3D Errors")
plot_error_stats(wagn_daily_stats, 'WAGN Daily')

print("WAGN Weekly 2D/3D Errors")
plot_error_stats(wagn_weekly_stats, 'WAGN Weekly')

print("WAGN Monthly 2D/3D Errors")
plot_error_stats(wagn_monthly_stats, 'WAGN Monthly')

# %% [Notebook cell 27]
import seaborn as sns
import matplotlib.pyplot as plt

# Function to compute correlation and plot
def plot_correlation(df, station_name):
    # Select GNSS error components
    errors = df[['dn','de','du']].copy()

    # Compute Pearson correlation
    corr = errors.corr(method='pearson')
    print(f"\n{station_name} Pearson Correlation:\n", corr)

    # Plot correlation matrix
    plt.figure(figsize=(6,5))
    sns.heatmap(corr, annot=True, cmap='coolwarm', vmin=-1, vmax=1)
    plt.title(f'{station_name} GNSS Error Correlation Matrix')
    plt.show()

    # Pairwise scatter plots
    sns.pairplot(errors)
    plt.suptitle(f'{station_name} GNSS Error Scatter Plots', y=1.02)
    plt.show()

# ----- ADCN Correlation -----
plot_correlation(adcn, 'ADCN')

# ----- WAGN Correlation -----
plot_correlation(wagn, 'WAGN')

# %% [Notebook cell 29]
# Function to compute correlation with meteorological parameters and plot
def plot_gnss_meteo_correlation(df, station_name):
    # Select GNSS errors and meteorological columns
    columns = ['dn','de','du','Temp.Dry [°C]','RelHumidity [%]','Press.QFF [hPa]']
    data = df[columns].copy()

    # Compute Pearson correlation
    corr = data.corr(method='pearson')
    print(f"\n{station_name} GNSS vs Meteorology Pearson Correlation:\n", corr)

    # Heatmap for correlation
    plt.figure(figsize=(8,6))
    sns.heatmap(corr, annot=True, cmap='coolwarm', vmin=-1, vmax=1)
    plt.title(f'{station_name} GNSS Errors vs Meteorological Parameters')
    plt.show()

    # Scatter plots
    for error in ['dn','de','du']:
        for met in ['Temp.Dry [°C]','RelHumidity [%]','Press.QFF [hPa]']:
            plt.figure(figsize=(5,4))
            sns.scatterplot(x=df[met], y=df[error])
            plt.xlabel(met)
            plt.ylabel(error)
            plt.title(f'{station_name}: {error} vs {met}')
            plt.show()

# ----- ADCN GNSS vs Meteorology -----
plot_gnss_meteo_correlation(adcn, 'ADCN')

# ----- WAGN GNSS vs Meteorology -----
plot_gnss_meteo_correlation(wagn, 'WAGN')

# %% [Notebook cell 30]
import matplotlib.pyplot as plt
import numpy as np

# Data from the table
stations = ['ADCN', 'WAGN']
intervals = ['Daily', 'Weekly', 'Monthly']

max_2d = np.array([[11.00, 7.98, 7.17], [9.20, 6.77, 7.84]])
mean_2d = np.array([[2.97, 2.37, 2.57], [2.17, 2.12, 2.28]])
max_3d = np.array([[25.00, 18.21, 19.63], [114.88, 106.37, 99.86]])
mean_3d = np.array([[7.21, 6.15, 6.90], [53.45, 49.62, 50.60]])

x = np.arange(len(intervals))  # interval positions
width = 0.35  # bar width

fig, axes = plt.subplots(2, 1, figsize=(12, 10), sharex=True)

# Plot 2D errors
axes[0].bar(x - width/2, max_2d[0], width, label='ADCN Max 2D', color='skyblue')
axes[0].bar(x + width/2, max_2d[1], width, label='WAGN Max 2D', color='lightcoral')
axes[0].plot(x, mean_2d[0], 'o--', color='blue', label='ADCN Mean 2D')
axes[0].plot(x, mean_2d[1], 'o--', color='red', label='WAGN Mean 2D')
axes[0].set_ylabel('2D Error (mm)')
axes[0].set_title('2D GNSS Errors by Interval and Station')
axes[0].legend()
axes[0].grid(True, linestyle='--', alpha=0.5)

# Plot 3D errors
axes[1].bar(x - width/2, max_3d[0], width, label='ADCN Max 3D', color='skyblue')
axes[1].bar(x + width/2, max_3d[1], width, label='WAGN Max 3D', color='lightcoral')
axes[1].plot(x, mean_3d[0], 'o--', color='blue', label='ADCN Mean 3D')
axes[1].plot(x, mean_3d[1], 'o--', color='red', label='WAGN Mean 3D')
axes[1].set_ylabel('3D Error (mm)')
axes[1].set_title('3D GNSS Errors by Interval and Station')
axes[1].set_xticks(x)
axes[1].set_xticklabels(intervals)
axes[1].legend()
axes[1].grid(True, linestyle='--', alpha=0.5)

plt.tight_layout()
plt.show()

# %% [Notebook cell 32]
import pandas as pd

ADCN = pd.read_excel("/content/ADCN.xlsx")
print(ADCN.columns)

# %% [Notebook cell 33]
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt

# Load Excel files
ADCN = pd.read_excel("/content/ADCN.xlsx")
WAGN = pd.read_excel("/content/WAGN  .xlsx")

# Strip spaces from column names
ADCN.columns = ADCN.columns.str.strip()
WAGN.columns = WAGN.columns.str.strip()

# Rename WAGN GNSS and meteorological columns to match ADCN
WAGN.rename(columns={
    'd North (mm)': 'dn',
    'd East (mm)': 'de',
    'd Up (mm)': 'du',
    'Temp.Dry [°C]': 'Temp',
    'RelHumidity [%]': 'Humidity',
    'Press.QFF [hPa]': 'Pressure'
}, inplace=True)

ADCN.rename(columns={
    'Temp.Dry [°C]': 'Temp',
    'RelHumidity [%]': 'Humidity',
    'Press.QFF [hPa]': 'Pressure'
}, inplace=True)

# Define GNSS and meteorological columns
gnss = ['dn', 'de', 'du']
meteo = ['Temp', 'Humidity', 'Pressure']

# Compute correlation and plot heatmaps
for df_name, df in zip(['ADCN', 'WAGN'], [ADCN, WAGN]):
    print(f"\n--- {df_name} GNSS vs Meteorological Correlation ---")
    corr = df[gnss + meteo].corr().loc[gnss, meteo]
    print(corr)

    plt.figure(figsize=(8,6))
    sns.heatmap(corr, annot=True, cmap='coolwarm', fmt=".2f")
    plt.title(f"{df_name}: GNSS vs Meteorological Parameters Correlation")
    plt.show()

# %% [Notebook cell 34]
# Step X: Export Results to Excel
output_file = "GNSS_Error_Analysis.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    # Save ADCN tables
    adcn_daily.to_excel(writer, sheet_name="ADCN_Daily")
    adcn_weekly.to_excel(writer, sheet_name="ADCN_Weekly")
    adcn_monthly.to_excel(writer, sheet_name="ADCN_Monthly")

    # Save WAGN tables
    wagn_daily.to_excel(writer, sheet_name="WAGN_Daily")
    wagn_weekly.to_excel(writer, sheet_name="WAGN_Weekly")
    wagn_monthly.to_excel(writer, sheet_name="WAGN_Monthly")

print(f"✅ All results exported successfully to {output_file}")

